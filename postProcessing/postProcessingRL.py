import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import seaborn as sns
from dataclasses import dataclass #, field
import fnmatch
import json
from qecc.polynomialCodes import generateBicycleCode
from qecc.codeUtilities import sameCode
#import copy 
import bisect
from openpyxl import load_workbook
from operator import attrgetter
from qecc.polynomialCodes import generateABmatrices, bicycleCodeFromAB
from qecc.logicals import calculateCodeDimension
import shutil
#from qecc.utils import NAMED_ERROR_RANGES
from matplotlib.transforms import Bbox  
import textwrap
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)


FONT_SIZE = 12
TICKS_FONT_SIZE = 10
SUMMARY_FONT_SIZE = 10
FIGURE_TITLE_FONT_SIZE = 14
TEST_ERROR_RANGE = np.linspace(10**-4, 10**-1, 5)


from qecc.utils import baselines
#allSizes = ((6,6), (9,6),(12,6),(15,3),
#            (21,18), (3,27),(5,15))

"""
Geometric5 reward with error range normalization:
Reward for code 108_8_10 was calculated as 0.5455695787963828 
Reward for code 144_12_12 was calculated as 0.5546604878872918 
Reward for code 288_12_18 was calculated as 0.6001150333418372 
Reward for code 360_12_24 was calculated as 0.6092059424327464 
Reward for code 72_12_6 was calculated as 0.3460360942686508 
Reward for code 756_16_34 was calculated as 0.6546604878872919 
Reward for code 90_8_10 was calculated as 0.506331144559866 


linear5 reward with and without normalization
Reward for code 108_8_10 is 0.0374625 engineered reward is: 0.375
Reward for code 144_12_12 is 0.037462499999999996 engineered reward is: 0.37499999999999994
Reward for code 288_12_18 is 0.041958 engineered reward is: 0.42
Reward for code 360_12_24 is 0.03996 engineered reward is: 0.4
Reward for code 72_12_6 is 0.035964 engineered reward is: 0.36000000000000004
Reward for code 756_16_34 is 0.0384615 engineered reward is: 0.385
Reward for code 90_8_10 is 0.035964 engineered reward is: 0.36000000000000004

[[n,k,d]] NetEncoding
Rater ℓ,m A B
[[72,12,6]] 1/12 6,6 x3+y+y2 y3+x+x2 Reward for code 72_12_6 is 0.035964 engineered reward is: 0.36000000000000004
[[90,8,10]] 1/23 15,3 x9+y+y2 1+x2+x7 Reward for code 90_8_10 is 0.035964 engineered reward is: 0.36000000000000004
[[108,8,10]] 1/27 9,6 x3+y+y2 y3+x+x2 Reward for code 108_8_10 is 0.0374625 engineered reward is: 0.375
[[144,12,12]] 1/24 12,6 x3+y+y2 y3+x+x2 Reward for code 144_12_12 is 0.037462499999999996 engineered reward is: 0.37499999999999994
[[288,12,18]] 1/48 12,12 x3+y2+y7 y3+x+x2 Reward for code 288_12_18 is 0.041958 engineered reward is: 0.42
[[360,12,≤24]] 1/60 30,6 x9+y+y2 y3+x25+x26 Reward for code 360_12_24 is 0.03996 engineered reward is: 0.4
[[756,16,≤34]] 1/95 21,18 x3+y10+y17 y5+x3+x1 Reward for code 756_16_34 is 0.0384615 engineered reward is: 0.385

  ┌────────────────┬───────────┬─────────┬────────┬───────────┬────────────┐
  │  [[n, k, d]]  │  Rate r   │ d_circ  │  p_0   │  p_L(p1)  │  p_L(p2)  │
  ├────────────────┼───────────┼─────────┼────────┼───────────┼────────────┤
  │ [[72,  12,  6]]│   1/12    │   ≤6    │ 0.0048 │  7×10⁻⁵  │  7×10⁻⁸  │
  │ [[90,   8, 10]]│   1/23    │   ≤8    │ 0.0053 │  5×10⁻⁶  │  4×10⁻¹⁰ │
  │ [[108,  8, 10]]│   1/27    │   ≤8    │ 0.0058 │  3×10⁻⁶  │  1×10⁻¹⁰ │
  │ [[144, 12, 12]]│   1/24    │   ≤10   │ 0.0065 │  2×10⁻⁷  │  8×10⁻¹³ │
  │ [[288, 12, 18]]│   1/48    │   ≤18   │ 0.0069 │  2×10⁻¹² │  1×10⁻²² │
  └────────────────┴───────────┴─────────┴────────┴───────────┴────────────┘
"""



@dataclass
class VisitedCode:
    l: int
    m: int
    aX: np.ndarray
    bX: np.ndarray
    aY: np.ndarray
    bY: np.ndarray
    #weightsX: np.ndarray
    #weightsZ: np.ndarray
    Hx: np.ndarray 
    Hz: np.ndarray 
    reward: float
    

def findWorstCodes(pathToCodeLogs, numberOfWorstCodesToGet = 100):
    aggregatedCodes = []  # bits tuple -> [counts (5,), samples, k]
    worstRewards = []
    for dirpath, _dirnames, filenames in os.walk(pathToCodeLogs):
        for name in fnmatch.filter(filenames, "codeEvaluations_*.jsonl"):
            with open(os.path.join(dirpath, name)) as fid:
                for line in fid:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        record = json.loads(line)
                        aX = record["aX"]
                        bX = record["bX"]
                        aY = record["aY"]
                        bY = record["bY"]
                        l = int(record["env_l"])
                        m = int(record["env_m"])
                        reward = float(record["reward"])
                        combinedErrorRate = np.asarray(record["logicalErrorCounts"], dtype=np.int64) + \
                                    np.asarray(record["decoderFailureCounts"], dtype=np.int64)
                        logicalQubits = int(record["numberOfLogicalQubits"])
                        Hx,Hz = generateBicycleCode(l,m, aX, aY, bX, bY)
                        #wX = codeWeights(Hx)
                        #wZ = codeWeights(Hz)
                        #code = VisitedCode(l,m,aX,bX,aY,bY,wX, wZ, Hx, Hz, reward)
                        code = VisitedCode(l,m,aX,bX,aY,bY, Hx, Hz, reward)
                        # Check if the new code is the same as a previously logged code:
                        if len(aggregatedCodes) < numberOfWorstCodesToGet: # Populate the list
                            bisect.insort(aggregatedCodes, code, key=attrgetter("reward"))
                            
                            
                        else: # The list is full of codes, check if this code is different
                            if not np.any([sameCode([code.Hx,code.Hz],[c.Hx,c.Hz]) for c in aggregatedCodes]): # The code is different than all codes in the list, check if its a worse performing code
                                if code.reward < aggregatedCodes[-1].reward:
                                    bisect.insort(aggregatedCodes, code, key=attrgetter("reward"))
                        if len(aggregatedCodes) > numberOfWorstCodesToGet:
                            aggregatedCodes = aggregatedCodes[0:numberOfWorstCodesToGet]
                    except json.JSONDecodeError:
                        continue  # tolerate a truncated last line from a killed run
    return aggregatedCodes
                    
                                
                                
        
    
    
    
    

# def policyEvaluation(pathToCrawl, worstCodesForReset):   
#     """
#     At this point we have some policies, and we want to evaluate them.
#     We also have some saved codes which were logged.
#     This postprocessing script does the following:
#         a.	crawls over the save policy data and locates pth files.
#         b.	loads saved policy from pth, and 
#         c.	deterministically evaluates the policy for ~ 2 * (l + m ) iterations. 
#         d.	Optional: Resets the environment to something other than 0 and repeats (say 100 times).
#         e.	Change the decoder type and rollout again.

#     """
#     # Crawl (walk) through the folder 
#     for dirpath, _dirnames, filenames in os.walk(pathToCrawl):
#         if "policy_weights.pth" not in filenames or "experiment.txt" not in filenames:
#             # skip if you can't find saved weights or an experiment log.
#             continue
        
#         policyWeightsPath = os.path.join(dirpath, "policy_weights.pth")
#         experimentPath = os.path.join(dirpath, "experiment.txt")
#         # First get the comments from the log:
#         comments = readComments(experimentPath)
#         useDict = comments.get("env_use_dict_observation", "False").lower() == "true"
#         if not useDict:
#             continue
#         l = int(comments["env_l"])
#         m = int(comments["env_m"])



    
#     l, m = run["l"], run["m"]
#     bitFlipping = run["comments"].get("env_bit_flipping", "False").lower() == "true"
#     net, kMin = buildPolicyNet(run, qeccDataRoot)
#     maxSteps = rolloutMultiplier * max(l, m)
#     visited = []

#     env = makeEnv(l, m, kMin, bitFlipping, numberOfSamples=1,
#                   numberOfIterations=numberOfIterations, errorRange=FIVE_POINT_GRID)
#     obs, _ = env.reset(seed=0)
#     zeroVisited, zeroCycle = rollout(net, env, l, m, maxSteps, obs)
#     for step, state in enumerate(zeroVisited):
#         aX, bX, aY, bY = state
#         visited.append(VisitedCode(l, m, np.array(aX), np.array(bX), np.array(aY),
#                                    np.array(bY), "zero", step, run["name"]))

#     try:
#         worst = worstLoggedCodes(qeccDataRoot, l, m, kMin, numWorst) if numWorst > 0 else []
#     except (ValueError, FileNotFoundError) as error:
#         print(f"  item (d) skipped for {run['name']}: no usable logged codes ({error})")
#         worst = []
#     for worstIndex, (aX, bX, aY, bY, _reward) in enumerate(worst):
#         env.reset(seed=worstIndex + 1)
#         startObs = seedEnvWithCode(env, aX, bX, aY, bY)
#         worstVisited, _cycle = rollout(net, env, l, m, maxSteps, startObs)
#         for step, state in enumerate(worstVisited):
#             vaX, vbX, vaY, vbY = state
#             visited.append(VisitedCode(l, m, np.array(vaX), np.array(vbX), np.array(vaY),
#                                        np.array(vbY), f"worst-{worstIndex}", step, run["name"]))
#     return visited, zeroCycle







def clipLine(s, maxLen=42):
    # if len(s) <= maxLen:
    #     returnedList = [s]
    # else:
    #     returnedList = [s[:maxLen], s[maxLen:]] # Ad-hoc solution, but realistically that's all I need
    return textwrap.wrap(s, maxLen) or [""]

def readComments(filePath):
    """Parse the '# key = value' comment header of an experiment.txt into a dict of strings.
    Returns {} if the file has no comment header (older runs have none)."""
    comments = {}
    with open(filePath) as fh:
        for line in fh:
            if not line.startswith("#"):
                continue                        # data rows have no leading '#'
            body = line.lstrip("#").strip()     # "# env_l = 9" -> "env_l = 9"
            if "=" in body:
                key, _, value = body.partition("=")
                comments[key.strip()] = value.strip()   # {"env_l": "9", ...}
    return comments

def markUnfreeze(a, xData, yData, label="encoder unfrozen", boxAt=(0.5, 0.88)):
    a.annotate(
        label,
        xy=(xData, yData), xycoords="data",              # tip: the real data point at the unfreeze eval
        xytext=boxAt,      textcoords="axes fraction",   # box: a fixed clear spot (tweak per panel)
        ha="center", va="center", fontsize=FONT_SIZE, color="#7a3200",
        bbox=dict(boxstyle="round,pad=0.35", fc="#fff3e0", ec="#e08a3c", lw=1.2),
        arrowprops=dict(arrowstyle="-|>", color="#e08a3c", lw=1.3, shrinkB=6),
    )
def buildExcelRowFromComments(comments, folderPath):
    """Map a run's comments + folder path to {columnLetter: value}, matching the hand-built
    parameterSweep.xlsx table (data area columns H..W)."""
    def g(key):
        v = comments.get(key)
        return "" if v in (None, "None") else v

    surrogate = comments.get("model_surrogate_model_path")
    hasSurrogate = surrogate not in (None, "None", "")
    encoderShort = ""
    if hasSurrogate:
        parts = surrogate.replace("\\", "/").rstrip("/").split("/")
        encoderShort = "/".join(p for p in parts[-2:] if p)

    code = f"{comments['env_l']},{comments['env_m']}" if "env_l" in comments and "env_m" in comments else ""

    dictForRow = {
        "H": folderPath,                              # Job name / folder name
        "I": int(comments.get("seed_for_environment")) if comments.get("seed_for_environment") else None,               # Seed
        "J": code,                                    # Code
        "K": int(comments.get("env_minimum_number_of_qubits")) if comments.get("env_minimum_number_of_qubits") else None,       # Number of qubits
        "M": comments.get("model_architecture"),                 # Network architecture
        "N": "Yes" if hasSurrogate else "No",         # Use pretrained encoder
        "O": float(comments.get("encoder_lr_factor")) if comments.get("encoder_lr_factor") else None,                  # Encoder lr factor
        "P": float(comments.get("index_to_unfreeze_encoder_updates")) if comments.get("index_to_unfreeze_encoder_updates") else None,  # Freeze until ?
        "Q": encoderShort,                            # Encoder
        "R": float(comments.get("entropy_eps")) if comments.get("entropy_eps") else None,                        # Entropy eps
        "S": "Done",                                  # Running / processed
        # T (Is good?) and U (Comments) are left blank for me to fill by hand.
    }
    return dictForRow


def appendToParameterSweep(xlsxPath, comments, folderPath, pngPath, experimentPath, sheetName="allData"):
    """Append one run as a new row to the parameter-sweep workbook. Dedupes by folder path
    (column H); adds the V/W hyperlink columns if missing. Returns True if a row was written."""
    if not xlsxPath or not os.path.exists(xlsxPath):
        print(f"[sweep] workbook not found, skipping: {xlsxPath}")
        raise ValueError("Workbook not found")
    XL_FOLDER_COLUMN = 8
    XL_COMMENTS_COLUMN = 23
    workBook = load_workbook(xlsxPath)    
    workSheet = workBook[sheetName] #if sheetName else workBook.active

    if workSheet["V1"].value in (None, ""):
        workSheet["V1"] = "Plot"
    if workSheet["W1"].value in (None, ""):
        workSheet["W1"] = "Data"

    # Check if the folder already appears in the excel
    normFolder = os.path.normcase(os.path.abspath(folderPath))
    for r in range(2, workSheet.max_row + 1):
        h = workSheet.cell(row=r, column=XL_FOLDER_COLUMN).value
        if h and os.path.normcase(os.path.abspath(str(h))) == normFolder:
            print(f"[sweep] already logged (row {r}), skipping: {folderPath}")
            return False

    # Get number of last row
    last = 1
    for r in range(2, workSheet.max_row + 1):
        for col in range(XL_FOLDER_COLUMN, XL_COMMENTS_COLUMN + 1):
            if workSheet.cell(row=r, column=col).value not in (None, ""):
                last = r
                break

    target = last + 1
    for colLetter, value in buildExcelRowFromComments(comments, folderPath).items():
        workSheet[f"{colLetter}{target}"] = value

    if pngPath and os.path.exists(pngPath):
        workSheet[f"V{target}"] = f'=HYPERLINK("{os.path.abspath(pngPath)}","plot")'
    if experimentPath and os.path.exists(experimentPath):
        workSheet[f"W{target}"] = f'=HYPERLINK("{os.path.abspath(experimentPath)}","experiment.txt")'
    try:
        workBook.save(xlsxPath)
    except PermissionError:
        alt = xlsxPath.replace(".xlsx", "_autofill.xlsx")
        workBook.save(alt)
        print(f"[sweep] '{os.path.basename(xlsxPath)}' is open in Excel; wrote to "
              f"'{os.path.basename(alt)}' instead. Close Excel to append directly.")
        return True
    print(f"[sweep] appended row {target}: {folderPath}")
    return True

def parsePolynomial(cell):
    """'[1 0 0 0 0 0]' -> np.array([1, 0, 0, 0, 0, 0]) of ints."""
    return np.fromstring(str(cell).strip().strip("[]"), sep=" ", dtype=int)

def analyseEvaluation(filePath, baseline = None):
    
    comments = readComments(filePath)
    env_l = int(comments["env_l"]) if "env_l" in comments else None
    env_m = int(comments["env_m"]) if "env_m" in comments else None
    
    codeParameters = f" for code parameters (l={comments['env_l']}, m={comments['env_m']})" if "env_l" in comments else ""
    sns.set_theme()
    df = pd.read_csv(filePath, sep='\t', comment = "#")
    dfEvalNumber = df.groupby(["evaluation number"])
    
    height = 2
    width = 3
    rewardAxIndex = 0
    entropyAxIndex = 1
    stepsToBestAxIndex = 2
    #boxPlotAxIndex = 2
    numberOfLogicalQubitsAxIndex = 3

    
    

    fig, allAx = plt.subplots(height, width, figsize=(24, 10), sharex=True,
                                            gridspec_kw={"width_ratios": [4] + [6]*(width -1)})
    allAx = allAx.flatten()
    newAx = []
    gs = allAx[0].get_gridspec()
    for i,a in enumerate(allAx):
        if i % width == 0:
            a.remove()
        else:
            newAx.append(a)
    axText = fig.add_subplot(gs[:,0])
    axText.set_axis_off()
    ax = np.array(newAx)
    polynomials = [p for p in ("postAction_aX","postAction_bX","postAction_aY","postAction_bY") if p in df.columns]
    

    xmin = 0
    xmax = len(dfEvalNumber.groups)
    baselineNumberOfQubits = 0
    if baseline is None and (env_l,env_m) in baselines.keys(): # Baseline was not given by the user, so # check if there is already a baseline in the dictionary at the top of this module                    
        if "env_reward_engineering" in comments:
            if comments["env_reward_engineering"].lower() == "true": # Remember that the comments are strings !
                baseline = baselines[(env_l,env_m)][f"reward_{comments.get("env_error_range")}"] # ERROR !!! I need to make sure I get the correct errorRange !!!
                baselineNumberOfQubits = baselines[(env_l,env_m)]["number of logical qubits"]
            else:
                baseline = baselines[(env_l,env_m)]["reward"]
                baselineNumberOfQubits = baselines[(env_l,env_m)]["number of logical qubits"]
                
    if baseline is not None: # Wow this is bad coding ! - basically, if either the user gave baseline or the baseline was found in the dict at the top of the file ...
        ax[rewardAxIndex].hlines(baseline, xmin, xmax, colors = "green", linestyles = "dotted", label = f"Reward for baseline code ({env_l},{env_m})")
    
    if polynomials:
        for col in ["postAction_aX", "postAction_bX", "postAction_aY", "postAction_bY"]:
            df[col.replace("postAction_", "")] = df[col].apply(parsePolynomial)
        logicalQubits = []
        for _, row in df.iterrows():
            A, B = generateABmatrices(env_l, env_m, # 
                                        np.where(row["aX"] !=0)[0], 
                                        np.where(row["aY"] !=0)[0], 
                                        np.where(row["bX"] !=0)[0], 
                                        np.where(row["bY"] !=0)[0])
                    
            Hx, Hz = bicycleCodeFromAB(A, B)
            logicalQubits.append( calculateCodeDimension(Hx,Hz))
        df["Number of logical qubits"] = logicalQubits


    # Data validation ?
    #if "postAction_k" in df.columns:
    #    assert np.all(df["Number of logical qubits"] == df["postAction_k"])
    

    

    if "Number of logical qubits" in df.columns or polynomials:
        sns.lineplot(data = df, x="evaluation number", y = "Number of logical qubits", ax = ax[numberOfLogicalQubitsAxIndex], estimator="median", label = "Logical qubits median")
    
        bestIndex = df.groupby("evaluation number")["reward"].idxmax()   # first max on ties
        kOfBest = df.loc[bestIndex].set_index("evaluation number")["Number of logical qubits"]
        ax[numberOfLogicalQubitsAxIndex].plot(kOfBest.index, kOfBest.values,
                                              color="crimson", linewidth=1.6,
                                              label="Number of logical qubits for best performing code")
        ax[numberOfLogicalQubitsAxIndex].hlines(baselineNumberOfQubits, xmin, xmax, colors = "green", linestyles = "dotted", label = f"Number of logical qubits for baseline code ({env_l},{env_m})")
        ax[numberOfLogicalQubitsAxIndex].legend(fontsize=FONT_SIZE)
    
    
    sns.lineplot(data=df, x= "evaluation number", y="reward", ax = ax[rewardAxIndex], label = "Average reward")
    maxReward = dfEvalNumber["reward"].max()
    ax[rewardAxIndex].plot(maxReward.index, maxReward.values, color="crimson", linewidth=1.6, label="Reward for best performing code")
    stepsToBest = (df.groupby("evaluation number")["reward"]
                 .apply(lambda s: int(np.argmax(s.to_numpy()))))   # argmax -> first max on ties
    ax[stepsToBestAxIndex].plot(stepsToBest.index, stepsToBest.values, marker="o", color = "crimson", label = "Minimum steps to best performing code")
    ax[stepsToBestAxIndex].set_title("Steps to best performing code per evaluation", fontsize=FONT_SIZE)
    ax[stepsToBestAxIndex].set_xlabel("Evaluation number", fontsize=FONT_SIZE)
    ax[stepsToBestAxIndex].set_ylabel("Step index of best (first max-reward) code", fontsize=FONT_SIZE)
    
    
    

    
    entropyColumns = [c for c in ("policy entropy",
                                "policy entropy aX", "policy entropy bX",
                                "policy entropy aY", "policy entropy bY")
                    if c in df.columns]

    if entropyColumns:
        meanByEval = dfEvalNumber[entropyColumns].mean()          # DataFrame: index=eval number, one col each
        for col in entropyColumns:
            isTotal = (col == "policy entropy")
            meanByEval[col].plot(
                ax=ax[entropyAxIndex],
                label="total" if isTotal else col.replace("policy entropy ", ""),
                color="black" if isTotal else None,               # components use the colour cycle
                linewidth=2.4 if isTotal else 1.4,
                linestyle="-"  if isTotal else "--",
                zorder=3       if isTotal else 2,
            )
        # If there is information about the entropy epsilon in the comments, include it in the title of the entropy plot.
        entropyInformationString = f" Ent. eps. = {comments['entropy_eps']}" if "entropy_eps" in comments else ""
        titleString = f"Policy entropy (total and per polynomial block)." + entropyInformationString
        ax[entropyAxIndex].set_title(titleString, fontsize = FONT_SIZE)
        ax[entropyAxIndex].set_xlabel("Evaluation number", fontsize = FONT_SIZE)
        ax[entropyAxIndex].set_ylabel("Entropy", fontsize = FONT_SIZE)
        ax[entropyAxIndex].legend(fontsize=FONT_SIZE)
    if "Encoder freeze" in df.columns:
        frozen = df["Encoder freeze"].astype(str).str.lower() == "true"   # normalise either dtype
        unfrozenRows = df.loc[~frozen, "evaluation number"]
        unfreezeEval = unfrozenRows.min() if len(unfrozenRows) else None    # None if never unfrozen in this run
        
        
        # If there is information about the encoder weights freeze / unfreeze include it as annotation in the plot.
        if unfreezeEval is not None: 
            ax[entropyAxIndex].axvline(unfreezeEval, color="crimson", linestyle="--", linewidth=1.5)
            ax[entropyAxIndex].text(unfreezeEval, ax[entropyAxIndex].get_ylim()[1], "  Encoder unfreeze",
                color="crimson", va="top", ha="left", fontsize=FONT_SIZE)

            atUnfreeze = df["evaluation number"] == unfreezeEval
            rewardY  = df.loc[atUnfreeze, "reward"].mean()          # mean reward at that eval
            entropyY = df.loc[atUnfreeze, "policy entropy"].mean()  # mean total entropy at that eval
            evalOrder = sorted(df["evaluation number"].unique())
            pos = evalOrder.index(unfreezeEval)
            #markUnfreeze(ax[boxPLotAxIndex], pos,          rewardY)    # boxplot:  category-index x, reward y
            markUnfreeze(ax[rewardAxIndex], unfreezeEval, rewardY)    # lineplot: value x,          reward y
            markUnfreeze(ax[entropyAxIndex], unfreezeEval, entropyY) 


    evalOrder = sorted(df["evaluation number"].unique())
    step = max(1, round(len(evalOrder) / 10))          # aim for ~10 labels
    tickPos = list(range(0, len(evalOrder), step))     # 0, 5, 10, ...

    # Set the figure property for the plots of the undiscounted reward.  Remember that entropy gets its own treatment.
    for i in [rewardAxIndex,]:
        ax[i].set_title('Undiscounted reward as a function of evaluation number' + codeParameters, fontsize = FONT_SIZE)
        ax[i].set_ylabel('Reward', fontsize = FONT_SIZE)
        ax[i].set_xlabel('Evaluation number', fontsize = FONT_SIZE)
        ax[i].legend(fontsize=FONT_SIZE)

    for a in ax:
        a.set_xticks(tickPos)
        a.tick_params(labelbottom=True)
        a.set_xticklabels([evalOrder[i] for i in tickPos], fontsize = TICKS_FONT_SIZE)

    
    
    pathBreakdown = os.path.split(filePath)
    fig.suptitle(f"Evaluation summary {os.path.basename(pathBreakdown[0])}\n" , fontsize = FIGURE_TITLE_FONT_SIZE)

    
    text = [f"{k} = {v}" for k, v in sorted(comments.items(), key=lambda kv: not kv[0].startswith("env_"))] # put env_ variables first
    text.append(f"{os.path.basename(pathBreakdown[0])}")
    wrappedText = []
    for t in text:
        temp = clipLine(t)
        for tt in temp:
            wrappedText.append(tt)

    figureExplanatoryText = "\n".join(wrappedText)
    axText.text(0.0, 1.0, figureExplanatoryText, transform=axText.transAxes,
                va="top", ha="left", fontsize=SUMMARY_FONT_SIZE, family="monospace", clip_on=False)
    

    folderName = os.path.basename(pathBreakdown[0])
    plt.tight_layout()
    
    # Save output as a png image in the same folder
    imageName = os.path.join(pathBreakdown[0], f"postProcessing.png")    
    plt.savefig(fname = imageName, dpi = 100)

    # Save a copy with no commentary text
    axText.set_visible(False)
    fig.canvas.draw()
    renderer = fig.canvas.get_renderer()
    plotsBbox = Bbox.union([a.get_tightbbox(renderer) for a in ax]).transformed(fig.dpi_scale_trans.inverted())
    noCommentsImageName = os.path.join(pathBreakdown[0], f"no_comments_{folderName}.png")
    fig.savefig(noCommentsImageName, bbox_inches=plotsBbox, dpi = 100)
    axText.set_visible(True)
    plt.show()
    # Try appending the processed experiment to the list of experiments in the excel sheet
    try:
        
        appendToParameterSweep(
            os.path.join(os.environ.get("QECC_DATA"), "parameterSweep.xlsx"), #noqa
            comments,
            folderPath=pathBreakdown[0],
            pngPath=imageName,
            experimentPath=filePath,
        )
    except Exception as exc:
        print(f"[sweep] could not update spreadsheet: {exc}")
    
    return df



def crawl(dataFolder, baseline=None):
    # Non-interactive backend, set BEFORE importing anything that pulls in pyplot,
    # so analyseEvaluation's plt.show() becomes a no-op and the crawl runs unattended.
    import matplotlib
    matplotlib.use("Agg")
    processed = 0
    skipped = 0
    failed = 0

    for dirpath, dirnames, filenames in os.walk(dataFolder):
        experiments = fnmatch.filter(filenames, "*experiment.txt")
        if not experiments:
            continue

        for experiment in experiments:
            experimentPath = os.path.join(dirpath, experiment)
            try:
                print(f"PROCESS: {experimentPath}")
                analyseEvaluation(experimentPath, baseline=baseline)
                processed += 1
            except Exception as exc:
                print(f"FAILED:  {experimentPath}  ->  {exc}")
                failed += 1
            finally:
                plt.close("all")                        # free the figure before the next file

    print(f"\nDone. processed={processed}, skipped={skipped}, failed={failed}")




def pruneByEntropyEps(dataFolder, threshold=0.3, dryRun=True):
    """Delete run folders whose experiment.txt has entropy_eps >= threshold.
    dryRun=True only prints what would go. Returns the list of matched folders."""
    matched = []
    for dirpath, _dirnames, filenames in os.walk(dataFolder):
        for experiment in fnmatch.filter(filenames, "*experiment.txt"):
            try:
                comments = readComments(os.path.join(dirpath, experiment))
                eps = float(comments["entropy_eps"])
            except (KeyError, ValueError, OSError):
                continue                      # no key, unparseable, or unreadable -> keep
            if eps >= threshold:
                matched.append((dirpath, eps))
            break                             # one verdict per folder
    for folder, eps in matched:
        print(f"{'WOULD DELETE' if dryRun else 'DELETING'} (entropy_eps={eps}): {folder}")
        if not dryRun:
            shutil.rmtree(folder)
    print(f"\n{len(matched)} folders {'would be' if dryRun else ''} deleted.")
    return matched

if __name__ == "__main__":
    analyseEvaluation(r"C:\Users\Omer\rl-qecc-data\2026-07-27_21-50-20_734978\search_9_6_seed_2236067_experiment.txt")
    #dataFolder = os.environ.get("QECC_DATA")
    #crawl(dataFolder)
    



                          